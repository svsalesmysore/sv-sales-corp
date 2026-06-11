#!/usr/bin/env python3
"""Regenerate the admin stock-count Excel template from the current catalog.
Run after catalog changes:  python3 scripts/gen_stock_template.py
Output: public/SV-Sales-Stock-Count-Template.xlsx  (one row per sellable unit)."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

cat = json.load(open('src/data/lion-catalog.json'))
rows = []
for p in cat['products']:
    if p.get('variants'):
        for v in p['variants']:
            rows.append({'product': f"{p['name']} — {v['option']}", 'model': v.get('model') or v['option'],
                         'brand': p.get('brand',''), 'category': p['categoryId']})
    else:
        rows.append({'product': p['name'], 'model': p.get('model',''),
                     'brand': p.get('brand',''), 'category': p['categoryId']})
rows.sort(key=lambda r: (r['brand'], r['category'], r['product']))

wb = Workbook(); ws = wb.active; ws.title = "Stock Count"
DARK="0B1220"; LIGHT="F1F5F9"
hfont=Font(bold=True,color="FFFFFF",size=11); hfill=PatternFill("solid",fgColor=DARK)
thin=Side(style="thin",color="E2E8F0"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
ws.merge_cells('A1:E1'); ws['A1']="S V Sales Corporation — Stock Count Sheet"; ws['A1'].font=Font(bold=True,size=14,color=DARK)
ws.merge_cells('A2:E2')
ws['A2']=("Fill TODAY'S quantity in column E ('Stock Qty'). Leave blank or 0 if out of stock. "
          "Do NOT change columns A-D. Then in Admin -> Stock -> Import Excel/CSV, upload this file.")
ws['A2'].font=Font(size=9,italic=True,color="64748B"); ws['A2'].alignment=Alignment(wrap_text=True,vertical='center'); ws.row_dimensions[2].height=40
HDR=4
for c,h in enumerate(["Product","Model","Brand","Category","Stock Qty (today)"],1):
    cell=ws.cell(row=HDR,column=c,value=h); cell.font=hfont; cell.fill=hfill; cell.border=border
    cell.alignment=Alignment(horizontal='center' if c==5 else 'left',vertical='center')
for i,r in enumerate(rows):
    rr=HDR+1+i
    ws.cell(row=rr,column=1,value=r['product']).border=border
    m=ws.cell(row=rr,column=2,value=r['model']); m.border=border; m.font=Font(name="Consolas",size=10,color="64748B")
    ws.cell(row=rr,column=3,value=r['brand']).border=border
    ws.cell(row=rr,column=4,value=r['category']).border=border
    q=ws.cell(row=rr,column=5); q.border=border; q.alignment=Alignment(horizontal='center'); q.fill=PatternFill("solid",fgColor="FFF7ED")
    if i%2:
        for c in range(1,5): ws.cell(row=rr,column=c).fill=PatternFill("solid",fgColor=LIGHT)
dv=DataValidation(type="whole",operator="greaterThanOrEqual",formula1="0",allow_blank=True,
                  error="Enter a whole number (0 or more)",errorTitle="Invalid quantity")
ws.add_data_validation(dv); dv.add(f"E{HDR+1}:E{HDR+len(rows)}")
for col,w in zip("ABCDE",[52,14,12,26,18]): ws.column_dimensions[col].width=w
ws.freeze_panes=f"A{HDR+1}"; ws.sheet_view.showGridLines=False
wb.save("public/SV-Sales-Stock-Count-Template.xlsx")
print(f"wrote public/SV-Sales-Stock-Count-Template.xlsx with {len(rows)} rows")
